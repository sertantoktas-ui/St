# -*- coding: utf-8 -*-
"""
Endustriyel Toz Emme Sistemi Hesaplama Araci
SMT Europe / Siloma / Saf Teknik katalog tabanli
"""
import re
import math
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference

# ---------------------------------------------------------------------------
# Yardimci: Excel 365 dinamik fonksiyonlari icin _xlfn / _xlfn._xlws on eki
# ---------------------------------------------------------------------------
XLWS_FUNCS = {"FILTER", "SORT", "SORTBY"}
XLFN_FUNCS = {"XLOOKUP", "XMATCH", "UNIQUE", "SEQUENCE", "RANDARRAY",
              "IFS", "MAXIFS", "MINIFS", "SWITCH", "TEXTJOIN"}

_func_re = re.compile(r"(?<![A-Za-z0-9_.])([A-Za-z][A-Za-z0-9_.]*)\(")


def fx(formula: str) -> str:
    def repl(m):
        name = m.group(1)
        up = name.upper()
        if up in XLWS_FUNCS:
            return "_xlfn._xlws." + name + "("
        if up in XLFN_FUNCS:
            return "_xlfn." + name + "("
        return name + "("
    return "=" + _func_re.sub(repl, formula.lstrip("="))


# ---------------------------------------------------------------------------
# Stil yardimcilari
# ---------------------------------------------------------------------------
NAVY = "1F3864"
BLUE = "2E5395"
LBLUE = "DDEBF7"
GREEN = "C6E0B4"
RED = "F8CBAD"
YELLOW = "FFE699"
GREY = "F2F2F2"
WHITE = "FFFFFF"

thin = Side(style="thin", color="B7B7B7")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

title_font = Font(name="Calibri", size=16, bold=True, color=WHITE)
subtitle_font = Font(name="Calibri", size=10, italic=True, color="555555")
header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
label_font = Font(name="Calibri", size=10, bold=True, color="1F3864")
input_font = Font(name="Calibri", size=10, color="0033CC", bold=True)
calc_font = Font(name="Calibri", size=10, color="000000")
result_font = Font(name="Calibri", size=11, bold=True, color="1F3864")

title_fill = PatternFill("solid", fgColor=NAVY)
header_fill = PatternFill("solid", fgColor=BLUE)
input_fill = PatternFill("solid", fgColor="FFF2CC")
calc_fill = PatternFill("solid", fgColor=LBLUE)
section_fill = PatternFill("solid", fgColor=GREY)


def title_bar(ws, row, text, span=8, height=26):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = height
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = title_fill


def section(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(size=11, bold=True, color=NAVY)
    c.fill = section_fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = section_fill


def label_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = label_font
    return c


def input_cell(ws, row, col, value=None, numfmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = input_font
    c.fill = input_fill
    c.border = border_all
    c.alignment = Alignment(horizontal="center")
    if numfmt:
        c.number_format = numfmt
    return c


def calc_cell(ws, row, col, formula, numfmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=fx(formula) if formula.startswith("=") is False else fx(formula))
    c.font = result_font if bold else calc_font
    c.fill = calc_fill
    c.border = border_all
    c.alignment = Alignment(horizontal="center")
    if numfmt:
        c.number_format = numfmt
    return c


def header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border_all
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 32


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()
wb.remove(wb.active)

# ===========================================================================
# 1) ÜRÜN VERİTABANI  (önce olusturulur, diger sayfalar buna referans verir)
# ===========================================================================
db = wb.create_sheet("Urun Veritabani")
set_widths(db, [14, 14, 16, 10, 12, 12, 13, 13, 12, 12, 14, 22, 30])

title_bar(db, 1, "URUN VERITABANI  -  SMT Europe / Siloma / Saf Teknik", span=13, height=26)
db.merge_cells("A2:M2")
note = db.cell(row=2, column=1, value=(
    "NOT: Asagidaki teknik degerler (debi, basinc, guc, alan) ilgili marka model serileri esas alinarak "
    "muhendislik tahmini olarak olusturulmustur. Kataloglarin genel urun sayfalarinda detayli teknik "
    "deger (m3/h, Pa, kW) yayinlanmadigindan, kesin teklif oncesi SMT Europe / Siloma / Saf Teknik "
    "yetkili satis/teknik servisinden guncel datasheet degerleri ile bu tablo dogrulanmalidir."))
note.font = Font(size=9, italic=True, color="C00000")
note.alignment = Alignment(wrap_text=True, vertical="center")
db.row_dimensions[2].height = 30

headers = ["Marka", "Urun Kodu", "Urun Tipi", "Cap (mm)", "Debi (m3/h)", "Basinc (Pa)",
           "Motor Gucu (kW)", "Filtre Alani (m2)", "Fan Devri (rpm)", "Fan Verimi (%)",
           "Hava Tanki Hacmi (m3)", "Temizlik Sistemi", "Aciklama"]
HEADER_ROW = 4
header_row(db, HEADER_ROW, headers)

STD_MOTORS = [0.55, 0.75, 1.1, 1.5, 2.2, 3, 4, 5.5, 7.5, 11, 15, 18.5, 22, 30, 37, 45, 55, 75, 90, 110, 132, 160, 200]


def std_motor(p):
    for m in STD_MOTORS:
        if m >= p:
            return m
    return STD_MOTORS[-1]


# ---- Filtre urunleri ----
filtre_models = [
    ("Saf Teknik", "SF-312", 3000, "Mekanik Sarsma"),
    ("Saf Teknik", "SF-315", 4500, "Mekanik Sarsma"),
    ("Saf Teknik", "SF-318", 6000, "Mekanik Sarsma"),
    ("Saf Teknik", "SF-321", 8000, "Mekanik Sarsma"),
    ("Saf Teknik", "SF-324", 10000, "Mekanik Sarsma"),
    ("Saf Teknik", "KP-12", 1200, "Mekanik Sarsma"),
    ("Saf Teknik", "KP-20", 2000, "Mekanik Sarsma"),
    ("Saf Teknik", "KP-30", 3000, "Mekanik Sarsma"),
    ("Saf Teknik", "KP-40", 4000, "Mekanik Sarsma"),
    ("Saf Teknik", "KP-50", 5000, "Mekanik Sarsma"),
    ("Saf Teknik", "KP-60", 6000, "Mekanik Sarsma"),
    ("Saf Teknik", "BB-1521", 15000, "Jet Pulse (Basincli Hava)"),
    ("Saf Teknik", "BB-1524", 20000, "Jet Pulse (Basincli Hava)"),
    ("Saf Teknik", "BB-1526", 26000, "Jet Pulse (Basincli Hava)"),
    ("Siloma", "SPE-30000", 30000, "Jet Pulse (Basincli Hava)"),
    ("Siloma", "SPE-40000", 40000, "Jet Pulse (Basincli Hava)"),
    ("Siloma", "SPE-55000", 55000, "Jet Pulse (Basincli Hava)"),
    ("SMT Europe", "SNF-10", 4000, "Titresimli (Vibrasyon)"),
    ("SMT Europe", "MS-15", 5000, "Titresimli (Vibrasyon)"),
    ("SMT Europe", "MS-25", 8000, "Titresimli (Vibrasyon)"),
    ("SMT Europe", "MSP-20", 7000, "Jet Pulse (Basincli Hava)"),
    ("SMT Europe", "MSP-35", 12000, "Jet Pulse (Basincli Hava)"),
    ("SMT Europe", "SWF-30", 10000, "Jet Pulse (Basincli Hava)"),
]

filtre_rows = []
for marka, kod, debi, temizlik in filtre_models:
    v_filt = 0.025 if temizlik == "Mekanik Sarsma" else (0.030 if temizlik.startswith("Titres") else 0.038)
    alan = round(debi / 3600 / v_filt, 1)
    basinc = round(1400 + min(800, (debi / 55000) * 800))
    motor_calc = (debi / 3600) * basinc / (1000 * 0.6)
    motor = std_motor(motor_calc)
    tank = 0 if temizlik in ("Mekanik Sarsma", "Titresimli (Vibrasyon)") else round(min(2.5, 0.3 + debi / 40000), 1)
    aciklama = f"{kod} torba/kartus filtre unitesi - {temizlik.lower()} temizlikli"
    filtre_rows.append((marka, kod, "Filtre", None, debi, basinc, motor, alan, None, None, tank, temizlik, aciklama))

# ---- Fan (transfer fani) urunleri ----
fan_models = [
    ("Saf Teknik", "SF-213", 2000),
    ("Saf Teknik", "SF-214", 3000),
    ("Saf Teknik", "SF-215", 4000),
    ("Saf Teknik", "SF-216", 5500),
    ("Saf Teknik", "SF-217", 7000),
    ("Saf Teknik", "SF-218", 9000),
    ("Saf Teknik", "SF-219", 12000),
    ("Saf Teknik", "SF-2110", 16000),
    ("Saf Teknik", "SF-2111", 20000),
    ("Saf Teknik", "SF-2112", 26000),
    ("Siloma", "STF-218", 9000),
    ("Siloma", "STF-2110", 16000),
    ("Siloma", "STF-2112", 26000),
    ("Siloma", "STF-2115", 35000),
    ("SMT Europe", "SMT-TF-10", 5000),
    ("SMT Europe", "SMT-TF-20", 10000),
    ("SMT Europe", "SMT-TF-30", 18000),
]

fan_rows = []
for marka, kod, debi in fan_models:
    basinc = round(1500 + min(900, (debi / 35000) * 900))
    devri = 2900 if debi < 8000 else 1450
    verimi = round(min(78, 65 + debi / 4000))
    motor_calc = (debi / 3600) * basinc / (1000 * verimi / 100)
    motor = std_motor(motor_calc)
    cap = round(math.sqrt(4 * (debi / 3600) / (math.pi * 20)) * 1000 / 10) * 10
    aciklama = f"{kod} radyal (santrifuj) transfer fani"
    fan_rows.append((marka, kod, "Fan", cap, debi, basinc, motor, None, devri, verimi, None, "-", aciklama))

all_rows = filtre_rows + fan_rows
DATA_START = HEADER_ROW + 1
DATA_END = DATA_START + len(all_rows) - 1

for i, row in enumerate(all_rows):
    r = DATA_START + i
    for j, val in enumerate(row):
        c = db.cell(row=r, column=j + 1, value=val)
        c.border = border_all
        c.alignment = Alignment(horizontal="center")
        if j in (4, 5, 6, 7, 8, 9, 10):
            c.number_format = "#,##0.0"
    if i % 2 == 1:
        for j in range(1, len(row) + 1):
            db.cell(row=r, column=j).fill = PatternFill("solid", fgColor="F7F9FC")

db.freeze_panes = "A5"

db_table = Table(displayName="tblUrunDB", ref=f"A{HEADER_ROW}:M{DATA_END}")
db_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
db.add_table(db_table)

DB = "'Urun Veritabani'"
TIPI_COL = f"{DB}!$C${DATA_START}:$C${DATA_END}"
DEBI_COL = f"{DB}!$E${DATA_START}:$E${DATA_END}"
BASINC_COL = f"{DB}!$F${DATA_START}:$F${DATA_END}"
MOTOR_COL = f"{DB}!$G${DATA_START}:$G${DATA_END}"
ALAN_COL = f"{DB}!$H${DATA_START}:$H${DATA_END}"
DEVRI_COL = f"{DB}!$I${DATA_START}:$I${DATA_END}"
VERIM_COL = f"{DB}!$J${DATA_START}:$J${DATA_END}"
TANK_COL = f"{DB}!$K${DATA_START}:$K${DATA_END}"
TEMIZ_COL = f"{DB}!$L${DATA_START}:$L${DATA_END}"
ACIK_COL = f"{DB}!$M${DATA_START}:$M${DATA_END}"
KOD_COL = f"{DB}!$B${DATA_START}:$B${DATA_END}"
MARKA_COL = f"{DB}!$A${DATA_START}:$A${DATA_END}"
CAP_COL = f"{DB}!$D${DATA_START}:$D${DATA_END}"

# ===========================================================================
# 2) BORU CAPI SECIM TABLOSU
# ===========================================================================
bc = wb.create_sheet("Boru Capi Tablosu")
set_widths(bc, [8, 18, 16, 22])
title_bar(bc, 1, "STANDART BORU CAPI SECIM TABLOSU (mm)", span=4)
bc.merge_cells("A2:D2")
n2 = bc.cell(row=2, column=1, value="Sistemde kullanilan standart spiral/duz dikiş boru capi serisi. Hesap sayfalari bu listeden bir ust standart capi otomatik secer (XLOOKUP, esit-veya-buyuk esleme).")
n2.font = subtitle_font
n2.alignment = Alignment(wrap_text=True)
bc.row_dimensions[2].height = 28

BC_HEADER = 4
header_row(bc, BC_HEADER, ["Sira", "Standart Cap (mm)", "Kesit Alani (m2)", "Tavsiye Edilen Hiz Araligi (m/s)"])
STD_CAPLAR = [80, 100, 120, 125, 140, 150, 160, 180, 200, 220, 250, 280, 300, 315, 350,
              400, 450, 500, 560, 600, 630, 710, 800]
BC_START = BC_HEADER + 1
BC_END = BC_START + len(STD_CAPLAR) - 1
for i, d in enumerate(STD_CAPLAR):
    r = BC_START + i
    bc.cell(row=r, column=1, value=i + 1).alignment = Alignment(horizontal="center")
    c = bc.cell(row=r, column=2, value=d)
    c.alignment = Alignment(horizontal="center")
    c.font = Font(bold=True)
    cell = bc.cell(row=r, column=3, value=fx(f"=PI()/4*(B{r}/1000)^2"))
    cell.number_format = "0.0000"
    cell.alignment = Alignment(horizontal="center")
    bc.cell(row=r, column=4, value="18 - 28 m/s (ahsap tozu tasima)").alignment = Alignment(horizontal="center")
    for col in range(1, 5):
        bc.cell(row=r, column=col).border = border_all
        if i % 2 == 1:
            bc.cell(row=r, column=col).fill = PatternFill("solid", fgColor="F7F9FC")

BC = "'Boru Capi Tablosu'"
STD_CAP_RANGE = f"{BC}!$B${BC_START}:$B${BC_END}"

# ===========================================================================
# 3) MAKINE GIRIS SAYFASI
# ===========================================================================
mg = wb.create_sheet("Makine Giris")
set_widths(mg, [4, 22, 18, 9, 10, 12, 14, 12, 11, 13, 13, 16])
title_bar(mg, 1, "MAKINE GIRIS SAYFASI", span=12)

mg.cell(row=2, column=1, value="Proje Adi:").font = label_font
input_cell(mg, 2, 2, "")
mg.cell(row=2, column=4, value="Tarih:").font = label_font
input_cell(mg, 2, 5, None, "dd.mm.yyyy")
mg.cell(row=2, column=7, value="Hazirlayan:").font = label_font
input_cell(mg, 2, 8, "")
mg.merge_cells("I2:L2")

MG_HEADER = 4
mg_headers = ["No", "Makine Adi", "Makine Tipi", "Makine\nAdedi", "Emis\nNoktasi Sayisi",
              "Emis Noktasi\nCapi (mm)", "Makine Hava\nDebisi (m3/h)", "Eszamanli\nCalisma K.(%)",
              "Calisma\nSuresi (sa/gun)", "Toplam Debi\n(m3/h)", "Gunluk Tuketim\n(m3/gun)", "Durum"]
header_row(mg, MG_HEADER, mg_headers)

MG_START = MG_HEADER + 1
N_MACHINES = 20
MG_END = MG_START + N_MACHINES - 1
for i in range(N_MACHINES):
    r = MG_START + i
    mg.cell(row=r, column=1, value=i + 1).alignment = Alignment(horizontal="center")
    for col in (2, 3):
        input_cell(mg, r, col, "")
    for col in (4, 5, 6, 7):
        input_cell(mg, r, col, "", "#,##0")
    input_cell(mg, r, 8, 100, "0")
    input_cell(mg, r, 9, 8, "0")
    calc_cell(mg, r, 10, f"=IF(B{r}=\"\",\"\",D{r}*G{r}*(H{r}/100))", "#,##0")
    calc_cell(mg, r, 11, f"=IF(B{r}=\"\",\"\",J{r}*I{r})", "#,##0")
    calc_cell(mg, r, 12, f'=IF(B{r}="","Bos Satir",IF(OR(D{r}<=0,G{r}<=0,H{r}<=0,H{r}>100),"HATA: Kontrol Et","OK"))')
    mg.cell(row=r, column=1).border = border_all
    mg.row_dimensions[r].height = 18

TOTAL_ROW = MG_END + 1
mg.cell(row=TOTAL_ROW, column=2, value="TOPLAM").font = Font(bold=True, size=11)
mg.merge_cells(start_row=TOTAL_ROW, start_column=2, end_row=TOTAL_ROW, end_column=9)
mg.cell(row=TOTAL_ROW, column=2).alignment = Alignment(horizontal="right")
calc_cell(mg, TOTAL_ROW, 10, f"=SUM(J{MG_START}:J{MG_END})", "#,##0", bold=True)
calc_cell(mg, TOTAL_ROW, 11, f"=SUM(K{MG_START}:K{MG_END})", "#,##0", bold=True)
mg.row_dimensions[TOTAL_ROW].height = 20

# Veri dogrulama (hatali giris kontrolu)
tip_list = '"Planya,Kalinlik Makinesi,Serit Testere,Daire Testere,CNC Router,Zimpara Makinesi,Delik Makinesi,Pres,Diger"'
dv_tip = DataValidation(type="list", formula1=tip_list, allow_blank=True, showErrorMessage=True,
                         errorTitle="Gecersiz Tip", error="Listeden bir makine tipi seciniz.")
mg.add_data_validation(dv_tip)
dv_tip.add(f"C{MG_START}:C{MG_END}")

dv_pct = DataValidation(type="decimal", operator="between", formula1=0, formula2=100,
                         allow_blank=True, showErrorMessage=True, errorTitle="Gecersiz Deger",
                         error="Eszamanli calisma katsayisi 0-100 arasinda olmalidir.")
mg.add_data_validation(dv_pct)
dv_pct.add(f"H{MG_START}:H{MG_END}")

dv_pos = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0,
                         allow_blank=True, showErrorMessage=True, errorTitle="Gecersiz Deger",
                         error="Negatif deger girilemez.")
mg.add_data_validation(dv_pos)
dv_pos.add(f"D{MG_START}:G{MG_END}")
dv_pos.add(f"I{MG_START}:I{MG_END}")

# Kosullu bicimlendirme - Durum kolonu
mg.conditional_formatting.add(f"L{MG_START}:L{MG_END}",
    FormulaRule(formula=[f'ISNUMBER(SEARCH("HATA",L{MG_START}))'], fill=PatternFill("solid", fgColor=RED)))
mg.conditional_formatting.add(f"L{MG_START}:L{MG_END}",
    FormulaRule(formula=[f'L{MG_START}="OK"'], fill=PatternFill("solid", fgColor=GREEN)))
mg.freeze_panes = "B5"

MGS = "'Makine Giris'"

# ===========================================================================
# 4) BORULAMA HESAPLARI
# ===========================================================================
bh = wb.create_sheet("Borulama Hesaplari")
set_widths(bh, [34, 16, 4, 40, 16])
title_bar(bh, 1, "BORULAMA HESAPLARI", span=5)

section(bh, 3, "GIRIS VERILERI", span=2)
inputs = [
    ("Ana Hat Uzunlugu (m)", 60, "#,##0.0"),
    ("Bransman (Dal Hatti) Uzunlugu (m)", 25, "#,##0.0"),
    ("45 derece Dirsek Sayisi (adet)", 4, "0"),
    ("90 derece Dirsek Sayisi (adet)", 6, "0"),
    ("T Baglanti Sayisi (adet)", 5, "0"),
    ("Reduksiyon Sayisi (adet)", 3, "0"),
    ("Boru Malzemesi", "Galvanizli Sac", None),
    ("Hedef Tasima Hizi (m/s)", 22, "0.0"),
    ("Guvenlik Payi (%)", 15, "0"),
]
r = 4
input_rows = {}
for label, default, fmt in inputs:
    label_cell(bh, r, 1, label)
    input_cell(bh, r, 2, default, fmt)
    input_rows[label] = r
    r += 1

L_ANA, L_BRN, L_D45, L_D90, L_TBAG, L_RED, L_MALZ, L_HIZ, L_GUV = [
    input_rows[k] for k in (
        "Ana Hat Uzunlugu (m)", "Bransman (Dal Hatti) Uzunlugu (m)", "45 derece Dirsek Sayisi (adet)",
        "90 derece Dirsek Sayisi (adet)", "T Baglanti Sayisi (adet)", "Reduksiyon Sayisi (adet)",
        "Boru Malzemesi", "Hedef Tasima Hizi (m/s)", "Guvenlik Payi (%)")]

dv_malz = DataValidation(type="list", formula1='"Galvanizli Sac,Paslanmaz Celik,Spiral Sac,PVC"',
                          allow_blank=False, showErrorMessage=True, errorTitle="Gecersiz Malzeme",
                          error="Listeden boru malzemesi seciniz.")
bh.add_data_validation(dv_malz)
dv_malz.add(f"B{L_MALZ}")

dv_hiz = DataValidation(type="decimal", operator="between", formula1=10, formula2=35,
                         allow_blank=True, showErrorMessage=True, errorTitle="Hiz Araligi Disi",
                         error="Hedef hiz 10-35 m/s araliginda olmalidir.")
bh.add_data_validation(dv_hiz)
dv_hiz.add(f"B{L_HIZ}")

# Malzeme surtunme katsayisi (lambda) referans tablosu
section(bh, 14, "MALZEME SURTUNME KATSAYISI (lambda)", span=2)
header_row(bh, 15, ["Boru Malzemesi", "Lambda"])
malz_tbl = [("Galvanizli Sac", 0.020), ("Paslanmaz Celik", 0.018), ("Spiral Sac", 0.022), ("PVC", 0.025)]
for i, (m, lam) in enumerate(malz_tbl):
    rr = 16 + i
    bh.cell(row=rr, column=1, value=m).border = border_all
    c = bh.cell(row=rr, column=2, value=lam)
    c.border = border_all
    c.number_format = "0.000"
    c.alignment = Alignment(horizontal="center")
MALZ_RANGE = f"$A$16:$A${15+len(malz_tbl)}"
LAMBDA_RANGE = f"$B$16:$B${15+len(malz_tbl)}"

# Yerel direnc katsayilari referans tablosu
section(bh, 22, "YEREL DIRENC (ZETA) KATSAYILARI", span=2)
header_row(bh, 23, ["Eleman", "Zeta"])
zeta_tbl = [("45 derece Dirsek", 0.15), ("90 derece Dirsek", 0.30), ("T Baglanti", 1.00), ("Reduksiyon", 0.20)]
for i, (m, z) in enumerate(zeta_tbl):
    rr = 24 + i
    bh.cell(row=rr, column=1, value=m).border = border_all
    c = bh.cell(row=rr, column=2, value=z)
    c.border = border_all
    c.number_format = "0.00"
    c.alignment = Alignment(horizontal="center")

# Hesap sonuclari
section(bh, 4, "HESAP SONUCLARI", span=2)
section(bh, 4, "", span=2)
hr = 5
calc_defs = []


def add_calc(row, label, formula, fmt=None, bold=False):
    label_cell(bh, row, 4, label)
    calc_cell(bh, row, 5, formula, fmt, bold)


add_calc(4, "Toplam Hava Debisi (m3/h)", f"=IFERROR({MGS}!J{TOTAL_ROW},0)", "#,##0")
add_calc(5, "Toplam Hava Debisi (m3/s)", "=E4/3600", "0.000")
add_calc(6, "Gerekli Boru Kesit Alani (m2)", f"=E5/B{L_HIZ}", "0.0000")
add_calc(7, "Gerekli Boru Capi (mm)", "=SQRT(4*E6/PI())*1000", "#,##0.0")
add_calc(8, "Onerilen Standart Cap (mm)", f"=XLOOKUP(E7,{STD_CAP_RANGE},{STD_CAP_RANGE},,1)", "#,##0", bold=True)
add_calc(9, "Secilen Borunun Kesit Alani (m2)", "=PI()/4*(E8/1000)^2", "0.0000")
add_calc(10, "Gercek Hava Hizi (m/s)", "=E5/E9", "0.00", bold=True)
add_calc(11, "Hava Yogunlugu (kg/m3)", "=1.2", "0.00")
add_calc(12, "Surtunme Katsayisi - lambda", f"=XLOOKUP(B{L_MALZ},{MALZ_RANGE},{LAMBDA_RANGE})", "0.000")
add_calc(13, "Toplam Hat Uzunlugu (m)", f"=B{L_ANA}+B{L_BRN}", "#,##0.0")
add_calc(14, "Statik (Surtunme) Basinc Kaybi (Pa)", f"=E12*(E13/(E8/1000))*(E11*E10^2/2)", "#,##0")
add_calc(15, "Yerel Direnc Katsayilari Toplami (sigma-zeta)",
         f"=B{L_D45}*0.15+B{L_D90}*0.3+B{L_TBAG}*1+B{L_RED}*0.2", "0.00")
add_calc(16, "Dinamik (Yerel) Basinc Kaybi (Pa)", "=E15*(E11*E10^2/2)", "#,##0")
add_calc(17, "Ara Toplam Basinc Kaybi (Pa)", "=E14+E16", "#,##0")
add_calc(18, "Guvenlik Payi Eklenmis Toplam Basinc Kaybi (Pa)", f"=E17*(1+B{L_GUV}/100)", "#,##0", bold=True)
label_cell(bh, 19, 4, "TOPLAM SISTEM BASINC KAYBI (Pa)")
bh.cell(row=19, column=4).font = Font(bold=True, size=12, color=NAVY)
c19 = calc_cell(bh, 19, 5, "=E18", "#,##0", bold=True)
c19.font = Font(bold=True, size=13, color=NAVY)
c19.fill = PatternFill("solid", fgColor=GREEN)

label_cell(bh, 20, 4, "Hiz Kontrolu / Uyari")
durum = bh.cell(row=20, column=5, value=fx(
    '=IF(E10<18,"UYARI: Hiz dusuk - boru icinde toz cokebilir",IF(E10>30,"UYARI: Hiz yuksek - asinma/enerji kaybi riski","Hiz uygun (18-28 m/s)"))'))
durum.font = Font(bold=True)
durum.alignment = Alignment(horizontal="center", wrap_text=True)
durum.border = border_all
bh.row_dimensions[20].height = 28

bh.conditional_formatting.add("E10", CellIsRule(operator="between", formula=["18", "28"], fill=PatternFill("solid", fgColor=GREEN)))
bh.conditional_formatting.add("E10", CellIsRule(operator="lessThan", formula=["18"], fill=PatternFill("solid", fgColor=RED)))
bh.conditional_formatting.add("E10", CellIsRule(operator="greaterThan", formula=["30"], fill=PatternFill("solid", fgColor=RED)))
bh.conditional_formatting.add("E10", CellIsRule(operator="between", formula=["28", "30"], fill=PatternFill("solid", fgColor=YELLOW)))
bh.conditional_formatting.add("E20", FormulaRule(formula=['ISNUMBER(SEARCH("UYARI",E20))'], fill=PatternFill("solid", fgColor=RED)))
bh.conditional_formatting.add("E20", FormulaRule(formula=['E20="Hiz uygun (18-28 m/s)"'], fill=PatternFill("solid", fgColor=GREEN)))

BHS = "'Borulama Hesaplari'"

# ===========================================================================
# 5) FILTRE SECIMI
# ===========================================================================
fs = wb.create_sheet("Filtre Secimi")
set_widths(fs, [34, 22, 4, 30, 20])
title_bar(fs, 1, "FILTRE SECIMI", span=5)

section(fs, 3, "SISTEM GEREKSINIMI", span=2)
label_cell(fs, 4, 1, "Toplam Hava Debisi (m3/h)")
calc_cell(fs, 4, 2, f"={BHS}!E4", "#,##0")

label_cell(fs, 5, 1, "Uygun Minimum Filtre Debisi (m3/h)")
SENTINEL = "999999999"
gerekli_filtre_debisi = (f'=MIN(FILTER({DEBI_COL},({TIPI_COL}="Filtre")*({DEBI_COL}>=B4),{SENTINEL}))')
calc_cell(fs, 5, 2, gerekli_filtre_debisi, "#,##0")

label_cell(fs, 6, 1, "Uygun Filtre Bulundu mu?")
calc_cell(fs, 6, 2, f'=IF(B5>={SENTINEL},"HAYIR - Veritabanini genisletin","EVET")', bold=True)

section(fs, 8, "SECILEN FILTRE OZELLIKLERI", span=2)
filtre_match = f'({TIPI_COL}="Filtre")*({DEBI_COL}=B5)'
flabels = [
    ("Marka", f"=XLOOKUP(1,{filtre_match},{MARKA_COL})", None),
    ("Filtre Modeli (Urun Kodu)", f"=XLOOKUP(1,{filtre_match},{KOD_COL})", None),
    ("Nominal Debi (m3/h)", "=B5", "#,##0"),
    ("Filtre Alani (m2)", f"=XLOOKUP(1,{filtre_match},{ALAN_COL})", "#,##0.0"),
    ("Fan Gucu (kW)", f"=XLOOKUP(1,{filtre_match},{MOTOR_COL})", "#,##0.0"),
    ("Hava Tanki Hacmi (m3)", f"=XLOOKUP(1,{filtre_match},{TANK_COL})", "#,##0.0"),
    ("Temizlik Sistemi", f"=XLOOKUP(1,{filtre_match},{TEMIZ_COL})", None),
    ("Aciklama", f"=XLOOKUP(1,{filtre_match},{ACIK_COL})", None),
]
frow = 9
for label, formula, fmt in flabels:
    label_cell(fs, frow, 1, label)
    calc_cell(fs, frow, 2, formula, fmt, bold=(label in ("Filtre Modeli (Urun Kodu)", "Marka")))
    frow += 1

section(fs, frow + 1, "MUHENDISLIK KONTROLU - FILTRE YUZEY HIZI (Air-to-Cloth Ratio)", span=2)
chk = frow + 2
label_cell(fs, chk, 1, "Filtre Yuzey Hizi (m/s) = (Debi/3600)/Alan")
calc_cell(fs, chk, 2, f"=(B5/3600)/B12", "0.000")
label_cell(fs, chk + 1, 1, "Kontrol (Tavsiye: 0.020 - 0.045 m/s)")
durum_fs = fs.cell(row=chk + 1, column=2, value=fx(
    f'=IF(B{chk}<0.02,"Filtre alani fazla buyuk (verimsiz)",IF(B{chk}>0.045,"UYARI: Filtre alani yetersiz olabilir","Uygun"))'))
durum_fs.font = Font(bold=True)
durum_fs.alignment = Alignment(horizontal="center", wrap_text=True)
durum_fs.border = border_all
durum_fs.fill = calc_fill

fs.conditional_formatting.add(f"B{chk+1}", FormulaRule(formula=['ISNUMBER(SEARCH("UYARI",B' + str(chk + 1) + '))'], fill=PatternFill("solid", fgColor=RED)))
fs.conditional_formatting.add(f"B{chk+1}", FormulaRule(formula=[f'B{chk+1}="Uygun"'], fill=PatternFill("solid", fgColor=GREEN)))
fs.conditional_formatting.add("B6", FormulaRule(formula=['ISNUMBER(SEARCH("HAYIR",B6))'], fill=PatternFill("solid", fgColor=RED)))
fs.conditional_formatting.add("B6", FormulaRule(formula=['B6="EVET"'], fill=PatternFill("solid", fgColor=GREEN)))

FS = "'Filtre Secimi'"
FILTRE_MODEL_ROW = 10
FILTRE_MARKA_ROW = 9
FILTRE_ALAN_ROW = 12
FILTRE_MOTOR_ROW = 13
FILTRE_TANK_ROW = 14
FILTRE_TEMIZ_ROW = 15
FILTRE_BASINC_FORMULA = f"=XLOOKUP(1,{filtre_match},{BASINC_COL})"
label_cell(fs, frow + 6, 1, "Filtre Basinc Kaybi (Pa) [Fan secimine eklenir]")
calc_cell(fs, frow + 6, 2, FILTRE_BASINC_FORMULA, "#,##0")
FILTRE_BASINC_ROW = frow + 6

# ===========================================================================
# 6) FAN SECIMI
# ===========================================================================
fn = wb.create_sheet("Fan Secimi")
set_widths(fn, [34, 22, 4, 30, 20])
title_bar(fn, 1, "FAN SECIMI", span=5)

section(fn, 3, "SISTEM GEREKSINIMI", span=2)
label_cell(fn, 4, 1, "Toplam Hava Debisi (m3/h)")
calc_cell(fn, 4, 2, f"={BHS}!E4", "#,##0")
label_cell(fn, 5, 1, "Borulama Basinc Kaybi (Pa)")
calc_cell(fn, 5, 2, f"={BHS}!E19", "#,##0")
label_cell(fn, 6, 1, "Filtre Basinc Kaybi (Pa)")
calc_cell(fn, 6, 2, f"={FS}!B{FILTRE_BASINC_ROW}", "#,##0")
label_cell(fn, 7, 1, "TOPLAM SISTEM BASINCI (Fan Secimi Icin) (Pa)")
c7 = calc_cell(fn, 7, 2, "=B5+B6", "#,##0", bold=True)
c7.fill = PatternFill("solid", fgColor=GREEN)

label_cell(fn, 9, 1, "Uygun Minimum Fan Debisi (m3/h)")
fan_cond = f'({TIPI_COL}="Fan")*({DEBI_COL}>=B4)*({BASINC_COL}>=B7)'
gerekli_fan_debisi = f"=MIN(FILTER({DEBI_COL},{fan_cond},{SENTINEL}))"
calc_cell(fn, 9, 2, gerekli_fan_debisi, "#,##0")

label_cell(fn, 10, 1, "Uygun Fan Bulundu mu?")
calc_cell(fn, 10, 2, f'=IF(B9>={SENTINEL},"HAYIR - Veritabanini genisletin","EVET")', bold=True)

section(fn, 12, "SECILEN FAN OZELLIKLERI", span=2)
fan_match = f'({TIPI_COL}="Fan")*({DEBI_COL}=B9)*({BASINC_COL}>=B7)'
fnlabels = [
    ("Marka", f"=XLOOKUP(1,{fan_match},{MARKA_COL})", None),
    ("Fan Modeli (Urun Kodu)", f"=XLOOKUP(1,{fan_match},{KOD_COL})", None),
    ("Cap (mm)", f"=XLOOKUP(1,{fan_match},{CAP_COL})", "#,##0"),
    ("Nominal Debi (m3/h)", "=B9", "#,##0"),
    ("Basinc Kapasitesi (Pa)", f"=XLOOKUP(1,{fan_match},{BASINC_COL})", "#,##0"),
    ("Veritabani Motor Gucu (kW)", f"=XLOOKUP(1,{fan_match},{MOTOR_COL})", "#,##0.0"),
    ("Fan Devri (rpm)", f"=XLOOKUP(1,{fan_match},{DEVRI_COL})", "#,##0"),
    ("Fan Verimi (%)", f"=XLOOKUP(1,{fan_match},{VERIM_COL})", "#,##0"),
]
fnrow = 13
for label, formula, fmt in fnlabels:
    label_cell(fn, fnrow, 1, label)
    calc_cell(fn, fnrow, 2, formula, fmt, bold=(label in ("Fan Modeli (Urun Kodu)", "Marka")))
    fnrow += 1

FAN_MARKA_ROW, FAN_MODEL_ROW, FAN_CAP_ROW, FAN_DEBI_ROW, FAN_BASINC_ROW, FAN_MOTORDB_ROW, FAN_DEVRI_ROW, FAN_VERIM_ROW = range(13, 21)

section(fn, fnrow + 1, "TEORIK MOTOR GUCU HESABI VE KONTROL", span=2)
tr = fnrow + 2
label_cell(fn, tr, 1, "Teorik Motor Gucu (kW) = Q[m3/s] x dP[Pa] / (1000 x Verim)")
calc_cell(fn, tr, 2, f"=(B4/3600)*B7/(1000*(B{FAN_VERIM_ROW}/100))", "0.00")
label_cell(fn, tr + 1, 1, "Kontrol (Teorik <= Veritabani Motoru x 1.1)")
durum_fn = fn.cell(row=tr + 1, column=2, value=fx(
    f'=IF(B{tr}>B{FAN_MOTORDB_ROW}*1.1,"UYARI: Motor gucu yetersiz olabilir","Uygun")'))
durum_fn.font = Font(bold=True)
durum_fn.border = border_all
durum_fn.fill = calc_fill
durum_fn.alignment = Alignment(horizontal="center")

section(fn, tr + 3, "ENERJI TUKETIMI", span=2)
er = tr + 4
label_cell(fn, er, 1, "Sistem Calisma Suresi (saat/gun)")
input_cell(fn, er, 2, 8, "0")
label_cell(fn, er + 1, 1, "Yillik Calisma Gunu (gun/yil)")
input_cell(fn, er + 1, 2, 250, "0")
label_cell(fn, er + 2, 1, "Motor Gucu (kW) [Kullanilan]")
calc_cell(fn, er + 2, 2, f"=MAX(B{tr},B{FAN_MOTORDB_ROW})", "0.00", bold=True)
label_cell(fn, er + 3, 1, "Enerji Tuketimi (kWh/gun)")
calc_cell(fn, er + 3, 2, f"=B{er+2}*B{er}", "#,##0.0")
label_cell(fn, er + 4, 1, "Enerji Tuketimi (kWh/yil)")
calc_cell(fn, er + 4, 2, f"=B{er+3}*B{er+1}", "#,##0")

ENERJI_KW_ROW = er + 2
ENERJI_GUN_ROW = er + 3
ENERJI_YIL_ROW = er + 4

fn.conditional_formatting.add(f"B{tr+1}", FormulaRule(formula=['ISNUMBER(SEARCH("UYARI",B' + str(tr + 1) + '))'], fill=PatternFill("solid", fgColor=RED)))
fn.conditional_formatting.add(f"B{tr+1}", FormulaRule(formula=[f'B{tr+1}="Uygun"'], fill=PatternFill("solid", fgColor=GREEN)))
fn.conditional_formatting.add("B10", FormulaRule(formula=['ISNUMBER(SEARCH("HAYIR",B10))'], fill=PatternFill("solid", fgColor=RED)))
fn.conditional_formatting.add("B10", FormulaRule(formula=['B10="EVET"'], fill=PatternFill("solid", fgColor=GREEN)))

FN = "'Fan Secimi'"

# ===========================================================================
# 7) TEKLIF OZETI
# ===========================================================================
tk = wb.create_sheet("Teklif Ozeti")
set_widths(tk, [30, 26, 4, 30, 22])
title_bar(tk, 1, "TOZ EMME SISTEMI - TEKLIF OZETI", span=5, height=30)

section(tk, 3, "MUSTERI / PROJE BILGILERI", span=5)
tk.cell(row=4, column=1, value="Musteri / Firma Adi:").font = label_font
input_cell(tk, 4, 2, "")
tk.cell(row=4, column=4, value="Teklif No:").font = label_font
input_cell(tk, 4, 5, "")
tk.cell(row=5, column=1, value="Proje:").font = label_font
input_cell(tk, 5, 2, f"={MGS}!B2")
tk.cell(row=5, column=4, value="Tarih:").font = label_font
input_cell(tk, 5, 5, f"={MGS}!E2", "dd.mm.yyyy")
tk.cell(row=6, column=1, value="Hazirlayan:").font = label_font
input_cell(tk, 6, 2, f"={MGS}!H2")

section(tk, 8, "SISTEM OZETI", span=5)
ozet = [
    ("Toplam Hava Debisi (m3/h)", f"={BHS}!E4", "#,##0"),
    ("Ana Boru Capi (mm)", f"={BHS}!E8", "#,##0"),
    ("Gercek Hava Hizi (m/s)", f"={BHS}!E10", "0.0"),
    ("Toplam Sistem Basinc Kaybi (Pa)", f"={FN}!B7", "#,##0"),
    ("Secilen Filtre Modeli", f"={FS}!B{FILTRE_MARKA_ROW}&\" - \"&{FS}!B{FILTRE_MODEL_ROW}", None),
    ("Filtre Alani (m2)", f"={FS}!B{FILTRE_ALAN_ROW}", "#,##0.0"),
    ("Secilen Fan Modeli", f"={FN}!B{FAN_MARKA_ROW}&\" - \"&{FN}!B{FAN_MODEL_ROW}", None),
    ("Fan Motor Gucu (kW)", f"={FN}!B{ENERJI_KW_ROW}", "0.0"),
    ("Tahmini Enerji Tuketimi (kWh/gun)", f"={FN}!B{ENERJI_GUN_ROW}", "#,##0.0"),
    ("Tahmini Enerji Tuketimi (kWh/yil)", f"={FN}!B{ENERJI_YIL_ROW}", "#,##0"),
]
orow = 9
for label, formula, fmt in ozet:
    label_cell(tk, orow, 1, label)
    cc = calc_cell(tk, orow, 2, formula, fmt, bold=True)
    cc.fill = PatternFill("solid", fgColor=LBLUE)
    orow += 1

section(tk, orow + 1, "KULLANILAN EKIPMAN LISTESI (BOM)", span=5)
bom_header = orow + 2
header_row(tk, bom_header, ["No", "Ekipman / Kalem", "Ozellik", "Miktar", "Birim"], start_col=1)
tk.merge_cells(start_row=bom_header, start_column=2, end_row=bom_header, end_column=2)
bom_rows = [
    ("Ana Boru Hatti", f"={BHS}!E8&\" mm Cap\"", f"={BHS}!B{L_ANA}", "m"),
    ("Bransman Hatlari", f"={BHS}!E8&\" mm Cap\"", f"={BHS}!B{L_BRN}", "m"),
    ("45 derece Dirsek", f"={BHS}!B{L_MALZ}", f"={BHS}!B{L_D45}", "adet"),
    ("90 derece Dirsek", f"={BHS}!B{L_MALZ}", f"={BHS}!B{L_D90}", "adet"),
    ("T Baglanti", f"={BHS}!B{L_MALZ}", f"={BHS}!B{L_TBAG}", "adet"),
    ("Reduksiyon", f"={BHS}!B{L_MALZ}", f"={BHS}!B{L_RED}", "adet"),
    ("Filtre Unitesi", f"={FS}!B{FILTRE_MARKA_ROW}&\" \"&{FS}!B{FILTRE_MODEL_ROW}", "1", "adet"),
    ("Fan (Aspirator)", f"={FN}!B{FAN_MARKA_ROW}&\" \"&{FN}!B{FAN_MODEL_ROW}", "1", "adet"),
]
br = bom_header + 1
for i, (kalem, ozellik, miktar, birim) in enumerate(bom_rows):
    rr = br + i
    tk.cell(row=rr, column=1, value=i + 1).alignment = Alignment(horizontal="center")
    tk.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=2)
    tk.cell(row=rr, column=2, value=kalem)
    tk.cell(row=rr, column=3, value=fx(ozellik) if isinstance(ozellik, str) and ozellik.startswith("=") else ozellik)
    tk.cell(row=rr, column=4, value=fx(miktar) if isinstance(miktar, str) and miktar.startswith("=") else miktar)
    tk.cell(row=rr, column=5, value=birim)
    for col in range(1, 6):
        tk.cell(row=rr, column=col).border = border_all
        tk.cell(row=rr, column=col).alignment = Alignment(horizontal="center" if col != 2 else "left", wrap_text=True)
    if i % 2 == 1:
        for col in range(1, 6):
            tk.cell(row=rr, column=col).fill = PatternFill("solid", fgColor="F7F9FC")

tk.merge_cells(start_row=br + len(bom_rows) + 1, start_column=1, end_row=br + len(bom_rows) + 1, end_column=5)
foot = tk.cell(row=br + len(bom_rows) + 1, column=1,
                value="Bu teklif ozeti otomatik hesaplama araci ile olusturulmustur. Kesin fiyatlandirma ve teknik onay icin yetkili satis muhendisi ile gorusulmesi tavsiye edilir.")
foot.font = Font(italic=True, size=9, color="777777")
foot.alignment = Alignment(wrap_text=True)

TK = "'Teklif Ozeti'"

# ===========================================================================
# 8) DASHBOARD
# ===========================================================================
da = wb.create_sheet("Dashboard")
wb.move_sheet("Dashboard", offset=-(len(wb.sheetnames) - 1))  # en basa tasi
set_widths(da, [26, 18, 4, 26, 18, 4, 26, 18])
title_bar(da, 1, "TOZ EMME SISTEMI - GOSTERGE PANELI (DASHBOARD)", span=8, height=30)

kpi_defs = [
    ("Toplam Hava Debisi", f"={BHS}!E4", "#,##0 \"m3/h\"", 1),
    ("Ana Boru Capi", f"={BHS}!E8", "#,##0 \"mm\"", 4),
    ("Gercek Hava Hizi", f"={BHS}!E10", "0.0 \"m/s\"", 7),
    ("Toplam Basinc Kaybi", f"={FN}!B7", "#,##0 \"Pa\"", 1),
    ("Secilen Filtre", f"={FS}!B{FILTRE_MODEL_ROW}", None, 4),
    ("Secilen Fan", f"={FN}!B{FAN_MODEL_ROW}", None, 7),
    ("Fan Motor Gucu", f"={FN}!B{ENERJI_KW_ROW}", "0.0 \"kW\"", 1),
    ("Yillik Enerji Tuketimi", f"={FN}!B{ENERJI_YIL_ROW}", "#,##0 \"kWh\"", 4),
    ("Filtre Alani", f"={FS}!B{FILTRE_ALAN_ROW}", "#,##0.0 \"m2\"", 7),
]
for i, (label, formula, fmt, col) in enumerate(kpi_defs):
    band_row = 3 + (i // 3) * 2
    da.cell(row=band_row, column=col, value=label).font = Font(bold=True, color="FFFFFF", size=10)
    da.cell(row=band_row, column=col).fill = PatternFill("solid", fgColor=BLUE)
    da.merge_cells(start_row=band_row, start_column=col, end_row=band_row, end_column=col + 1)
    da.cell(row=band_row, column=col).alignment = Alignment(horizontal="center")
    vcell = da.cell(row=band_row + 1, column=col, value=fx(formula))
    da.merge_cells(start_row=band_row + 1, start_column=col, end_row=band_row + 1, end_column=col + 1)
    vcell.font = Font(bold=True, size=15, color=NAVY)
    vcell.fill = PatternFill("solid", fgColor=LBLUE)
    vcell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        vcell.number_format = fmt
    da.row_dimensions[band_row + 1].height = 30
    for rr in (band_row, band_row + 1):
        da.cell(row=rr, column=col).border = border_all
        da.cell(row=rr, column=col + 1).border = border_all

KPI_LAST_ROW = 3 + ((len(kpi_defs) - 1) // 3) * 2 + 1  # son KPI bandinin deger satiri

section(da, KPI_LAST_ROW + 2, "BASINC KAYBI BILESENLERI", span=4)
PK_HEADER = KPI_LAST_ROW + 3
header_row(da, PK_HEADER, ["Bilesen", "Deger (Pa)"])
da.cell(row=PK_HEADER + 1, column=1, value="Statik (Surtunme) Kaybi")
da.cell(row=PK_HEADER + 1, column=2, value=fx(f"={BHS}!E14"))
da.cell(row=PK_HEADER + 2, column=1, value="Dinamik (Yerel) Kaybi")
da.cell(row=PK_HEADER + 2, column=2, value=fx(f"={BHS}!E16"))
da.cell(row=PK_HEADER + 3, column=1, value="Filtre Basinc Kaybi")
da.cell(row=PK_HEADER + 3, column=2, value=fx(f"={FS}!B{FILTRE_BASINC_ROW}"))
for rr in (PK_HEADER + 1, PK_HEADER + 2, PK_HEADER + 3):
    da.cell(row=rr, column=1).border = border_all
    da.cell(row=rr, column=2).border = border_all
    da.cell(row=rr, column=2).number_format = "#,##0"

chart = BarChart()
chart.title = "Basinc Kaybi Bilesenleri (Pa)"
chart.type = "col"
chart.y_axis.title = "Pa"
data = Reference(da, min_col=2, min_row=PK_HEADER, max_row=PK_HEADER + 3)
cats = Reference(da, min_col=1, min_row=PK_HEADER + 1, max_row=PK_HEADER + 3)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width = 11
chart.height = 7.5
da.add_chart(chart, f"D{PK_HEADER}")

CHART2_ROW = PK_HEADER + 16
section(da, CHART2_ROW, "MAKINE BAZLI HAVA DEBISI DAGILIMI", span=8)

chart2 = BarChart()
chart2.title = "Makine Bazli Toplam Debi (m3/h)"
chart2.type = "bar"
chart2.y_axis.title = "m3/h"
data2 = Reference(mg, min_col=10, min_row=MG_HEADER, max_row=MG_END)
cats2 = Reference(mg, min_col=2, min_row=MG_START, max_row=MG_END)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width = 16
chart2.height = 8
da.add_chart(chart2, f"A{CHART2_ROW + 1}")

da.sheet_view.showGridLines = False

# ===========================================================================
# Tanimli Adlar (Defined Names) - workbook seviyesinde
# ===========================================================================
def add_name(name, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=ref)


add_name("ToplamHavaDebisi", f"{BHS}!$E$4")
add_name("AnaBoruCapi", f"{BHS}!$E$8")
add_name("ToplamBasincKaybi", f"{FN}!$B$7")
add_name("StandartCaplar", STD_CAP_RANGE)
add_name("FiltreModeli", f"{FS}!$B${FILTRE_MODEL_ROW}")
add_name("FanModeli", f"{FN}!$B${FAN_MODEL_ROW}")
add_name("FanMotorGucu", f"{FN}!$B${ENERJI_KW_ROW}")
add_name("EnerjiYillik", f"{FN}!$B${ENERJI_YIL_ROW}")

# Sayfa sirasi: Dashboard ilk sirada olmali (yukarida tasindi), digerleri mantikli sirada
order = ["Dashboard", "Makine Giris", "Borulama Hesaplari", "Boru Capi Tablosu",
         "Filtre Secimi", "Fan Secimi", "Urun Veritabani", "Teklif Ozeti"]
wb._sheets = [wb[s] for s in order]
for ws in wb.worksheets:
    ws.sheet_view.zoomScale = 100

wb.active = 0
print("Step4 OK - tum sayfalar tamamlandi")
wb.save("/home/user/St/dust_collection_tool.xlsx")
